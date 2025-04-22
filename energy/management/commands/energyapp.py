import os
import logging
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from tqdm import tqdm
from energy.models import Country, EnergyData

# Set up logging
logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Load renewable energy data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='energy/data/database.xlsx',
            help='Path to Excel file (relative to project root)'
        )

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, options['file'])
        
        # Check if file exists
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        try:
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb['Data']
        except KeyError:
            self.stderr.write(self.style.ERROR("Worksheet 'Data' not found"))
            return
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error loading Excel file: {e}"))
            return

        # Get headers
        headers = [cell.value for cell in sheet[4]]
        required_columns = ['Country Name', 'Country Code', 'Type', 'Region', 'IncomeGroup']
        #print(headers)
        # Validate headers
        missing = [col for col in required_columns if col not in headers]
        if missing:
            self.stderr.write(self.style.ERROR(f"Missing required columns: {missing}"))
            return

        # Clear existing data
        Country.objects.all().delete()
        EnergyData.objects.all().delete()
        self.stdout.write(self.style.WARNING("Cleared existing data"))

        success_count = 0
        error_count = 0
        #year_columns = [str(year) for year in range(1990, 2016)]  # All years from 1990 to 2015
        year_columns = list(range(1990, 2016))
        # Read data rows
        rows = list(sheet.iter_rows(min_row=5, values_only=True))
        
        for row in tqdm(rows, desc="Importing data"):
            try:
                if not row[headers.index('Type')] or row[headers.index('Type')] != 'Country':
                    continue

                # Create or update Country
                country, _ = Country.objects.get_or_create(
                    name=row[headers.index('Country Name')] or 'Unknown',
                    defaults={
                        'code': row[headers.index('Country Code')] or None,
                        'type': row[headers.index('Type')] or 'Unknown',
                        'region': row[headers.index('Region')] or 'Unknown',
                        'income_group': row[headers.index('IncomeGroup')] or 'Unknown',
                        'renewable_share': float(row[headers.index(2015)] or 0.0)
                    }
                )

                # Create EnergyData for all years
                energy_data = []
                for year in year_columns:
                    if year in headers:
                        value = row[headers.index(year)]
                        try:
                            value = float(value) if value is not None else 0.0
                        except (ValueError, TypeError):
                            value = 0.0
                        energy_data.append(
                            EnergyData(
                                country=country,
                                year=int(year),
                                renewable_share=value
                            )
                        )
                print(energy_data)
                EnergyData.objects.bulk_create(energy_data, ignore_conflicts=True)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                logger.error(f"Error processing row {row[0]}: {e}")

        self.stdout.write(self.style.SUCCESS(
            f"Import completed! Success: {success_count}, Failed: {error_count}\n"
            f"Total countries: {Country.objects.count()}\n"
            f"Total energy records: {EnergyData.objects.count()}"
        ))